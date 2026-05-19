"""患者一括取り込み API"""
import csv
import io
import json
import re
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from app.database import get_db
from app.api.auth import require_staff
from app.models.patient import Patient
from app.schemas.patient import _normalize_phone, _normalize_name, build_name
from app.api.patients import _generate_patient_number, _normalize_for_compare

router = APIRouter(prefix="/api/patients/import", tags=["patient-import"])

# ─── 列名推定パターン ───────────────────────────────
FIELD_PATTERNS: dict[str, list[str]] = {
    "last_name":   ["姓", "苗字", "last_name", "family_name"],
    "middle_name": ["ミドルネーム", "middle_name", "middle"],
    "first_name":  ["名", "first_name", "given_name"],
    "reading":     ["読み方", "ふりがな", "フリガナ", "カナ", "reading", "kana"],
    "full_name":   ["氏名", "患者名", "お名前", "full_name"],
    "phone":       ["電話番号", "電話", "tel", "phone", "携帯", "連絡先"],
    "birth_date":  ["生年月日", "誕生日", "birthday", "birth_date"],
    "email":       ["メール", "メールアドレス", "email", "e-mail"],
    "notes":       ["備考", "メモ", "note", "notes", "コメント"],
}

# "名前" / "name" は曖昧 → 姓列有無で判定
AMBIGUOUS_NAME_HEADERS = ["名前", "name"]

STANDARD_FIELDS_SPLIT = [
    "last_name", "middle_name", "first_name",
    "reading", "phone", "birth_date", "email", "notes",
]
STANDARD_FIELDS_FULL = [
    "full_name", "reading", "phone", "birth_date", "email", "notes",
]

# ─── テンプレート列ラベル ──────────────────────────────
TEMPLATE_HEADERS_SPLIT = ["姓", "ミドルネーム", "名", "読み方", "電話番号", "生年月日", "メールアドレス", "備考"]
TEMPLATE_HEADERS_FULL = ["フルネーム", "読み方", "電話番号", "生年月日", "メールアドレス", "備考"]
TEMPLATE_SAMPLE_SPLIT = ["山田", "", "太郎", "ヤマダ タロウ", "090-1234-5678", "1990-01-15", "yamada@example.com", ""]
TEMPLATE_SAMPLE_FULL = ["山田 太郎", "ヤマダ タロウ", "090-1234-5678", "1990-01-15", "yamada@example.com", ""]


# ─── テンプレートダウンロード API ──────────────────────
@router.get("/template/csv")
async def download_template_csv():
    """通常モード向け CSV テンプレート（UTF-8 BOM 付き）"""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS_SPLIT)
    writer.writerow(TEMPLATE_SAMPLE_SPLIT)
    content = "\ufeff" + buf.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="patient_import_template.csv"'},
    )


@router.get("/template/xlsx")
async def download_template_xlsx():
    """Excel テンプレート（通常モード + フルネームモードの 2 シート）"""
    from openpyxl import Workbook
    wb = Workbook()

    # シート1: 通常モード
    ws_split = wb.active
    ws_split.title = "通常モード"
    ws_split.append(TEMPLATE_HEADERS_SPLIT)
    ws_split.append(TEMPLATE_SAMPLE_SPLIT)

    # シート2: フルネームモード
    ws_full = wb.create_sheet("フルネームモード")
    ws_full.append(TEMPLATE_HEADERS_FULL)
    ws_full.append(TEMPLATE_SAMPLE_FULL)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="patient_import_template.xlsx"'},
    )


# ─── ファイル解析 ────────────────────────────────────
def _read_file(contents: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """CSV / Excel からヘッダー行 + データ行を返す"""
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _read_xlsx(contents)
    elif lower.endswith(".csv"):
        return _read_csv(contents)
    else:
        raise HTTPException(400, "対応形式は .csv / .xlsx です")


def _read_csv(contents: bytes) -> tuple[list[str], list[list[str]]]:
    for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            text = contents.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        raise HTTPException(400, "ファイルの文字コードを判別できません")

    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)
    if len(rows_raw) < 2:
        raise HTTPException(400, "ヘッダー行とデータ行が必要です")

    headers = [c.strip() for c in rows_raw[0]]
    data = rows_raw[1:]
    return headers, data


def _read_xlsx(contents: bytes) -> tuple[list[str], list[list[str]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows_raw) < 2:
        raise HTTPException(400, "ヘッダー行とデータ行が必要です")

    headers = [str(c).strip() if c is not None else "" for c in rows_raw[0]]
    data = [
        [str(c).strip() if c is not None else "" for c in row]
        for row in rows_raw[1:]
    ]
    return headers, data


# ─── 自動マッピング ──────────────────────────────────
def suggest_mapping(headers: list[str]) -> tuple[dict[str, int], str]:
    """ヘッダー名から各標準項目への列インデックスを推定し、推定モードも返す"""
    mapping: dict[str, int] = {}
    used_indices: set[int] = set()
    h_lower = [h.lower().strip() for h in headers]

    # 1) 明確なパターンでマッチ
    for field, patterns in FIELD_PATTERNS.items():
        for idx, h in enumerate(h_lower):
            if idx in used_indices:
                continue
            if h in [p.lower() for p in patterns]:
                mapping[field] = idx
                used_indices.add(idx)
                break

    # 2) 曖昧な "名前" / "name" の処理
    for idx, h in enumerate(h_lower):
        if idx in used_indices:
            continue
        if h in [a.lower() for a in AMBIGUOUS_NAME_HEADERS]:
            if "last_name" in mapping:
                # 姓がある → これは名 (first_name)
                if "first_name" not in mapping:
                    mapping["first_name"] = idx
                    used_indices.add(idx)
            else:
                # 姓がない → フルネーム
                if "full_name" not in mapping:
                    mapping["full_name"] = idx
                    used_indices.add(idx)
            break

    # 3) モード判定
    if "last_name" in mapping and "first_name" in mapping:
        mode = "split"
    elif "full_name" in mapping:
        mode = "full_name"
    elif "last_name" in mapping:
        # 姓だけある → split として扱い、名は未割当
        mode = "split"
    else:
        mode = "full_name"

    return mapping, mode


# ─── データ変換ヘルパー ──────────────────────────────
def _is_empty_row(row: list[str]) -> bool:
    return all(not c.strip() for c in row)


def try_split_name(value: str) -> tuple[str, str] | None:
    """氏名1列を姓/名に分割を試みる。安全に分割できる場合のみ (姓, 名) を返す。

    ルール:
    - 半角/全角スペースを正規化して1つのスペースにする
    - ちょうど2トークンに分かれる場合のみ分割候補
    - 3トークン以上（外国人名等）やスペースなしは None
    """
    if not value or not value.strip():
        return None
    normalized = re.sub(r'[\s\u3000]+', ' ', value.strip())
    parts = normalized.split(' ')
    if len(parts) == 2 and parts[0] and parts[1]:
        return (parts[0], parts[1])
    return None


def _parse_date(val: str) -> date | None:
    """さまざまな日付形式を試して date に変換"""
    if not val or not val.strip():
        return None
    val = val.strip()
    # 全角数字→半角
    val = val.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    # 8桁数字 (YYYYMMDD)
    if len(val) == 8 and val.isdigit():
        try:
            return datetime.strptime(val, "%Y%m%d").date()
        except ValueError:
            pass
    raise ValueError(f"日付変換不可: {val}")


def _extract_row(row: list[str], mapping: dict[str, int]) -> dict[str, str]:
    """1行分のデータを mapping に従ってフィールド→値の dict にする"""
    result: dict[str, str] = {}
    for field, idx in mapping.items():
        if idx < len(row):
            result[field] = row[idx].strip()
        else:
            result[field] = ""
    return result


# ─── 重複チェック ─────────────────────────────────────
async def _find_duplicates(
    db: AsyncSession, name: str, phone: str | None, reading: str | None, birth: date | None,
) -> list[dict[str, Any]]:
    """既存患者との重複候補を返す"""
    conditions = []
    name_norm = _normalize_for_compare(name)
    name_nospace = name_norm.replace(" ", "")

    if name_nospace:
        conditions.append(
            func.lower(func.replace(func.replace(Patient.name, "\u3000", " "), " ", ""))
                .ilike(f"%{name_nospace}%")
        )

    phone_norm = _normalize_phone(phone)
    if phone_norm:
        conditions.append(
            func.replace(func.replace(Patient.phone, "-", ""), "ー", "") == phone_norm
        )

    if not conditions:
        return []

    result = await db.execute(
        select(Patient).where(
            Patient.is_active == True,  # noqa: E712
            or_(*conditions),
        ).limit(10)
    )
    candidates = result.scalars().all()
    if not candidates:
        return []

    dupes: list[dict[str, Any]] = []
    for p in candidates:
        reasons = []
        p_name = _normalize_for_compare(p.name).replace(" ", "")
        if name_nospace and p_name == name_nospace:
            reasons.append("氏名一致")
        if phone_norm and p.phone and _normalize_phone(p.phone) == phone_norm:
            reasons.append("電話番号一致")
        if reading and p.reading:
            if _normalize_for_compare(p.reading).replace(" ", "") == _normalize_for_compare(reading).replace(" ", ""):
                reasons.append("読み方一致")
        if birth and p.birth_date and p.birth_date == birth:
            reasons.append("生年月日一致")
        if reasons:
            dupes.append({
                "id": p.id,
                "name": p.name,
                "patient_number": p.patient_number,
                "phone": p.phone,
                "reading": p.reading,
                "birth_date": p.birth_date.isoformat() if p.birth_date else None,
                "reasons": reasons,
            })
    return dupes


# ─── Preview API ──────────────────────────────────────
@router.post("/preview")
async def import_preview(file: UploadFile = File(...), _auth: dict = Depends(require_staff)):
    """ファイルを解析し、列候補・推定マッピング・プレビューを返す"""
    contents = await file.read()
    if not contents:
        raise HTTPException(400, "ファイルが空です")

    headers, data = _read_file(contents, file.filename or "import.csv")
    mapping, mode = suggest_mapping(headers)

    # プレビュー行 (先頭10行、空行除く)
    preview: list[dict[str, str]] = []
    for row in data[:20]:
        if _is_empty_row(row):
            continue
        preview.append(_extract_row(row, mapping))
        if len(preview) >= 10:
            break

    # full_name モード時、サンプルデータから姓名分割の可否を判定
    splittable_hint = False
    if mode == "full_name" and preview:
        splittable_count = sum(
            1 for r in preview if try_split_name(r.get("full_name", "")) is not None
        )
        splittable_hint = splittable_count >= len(preview) * 0.5

    return {
        "columns": headers,
        "suggested_mapping": mapping,
        "suggested_mode": mode,
        "preview_rows": preview,
        "total_data_rows": len([r for r in data if not _is_empty_row(r)]),
        "splittable_hint": splittable_hint,
    }


# ─── Execute API ──────────────────────────────────────
# update_existing で更新を許可するフィールド（安全のため限定）
UPDATABLE_FIELDS = ["phone", "email", "reading", "notes"]


@router.post("/execute")
async def import_execute(
    file: UploadFile = File(...),
    mode: str = Form(...),
    mapping_json: str = Form(...),
    row_actions_json: str = Form(""),
    db: AsyncSession = Depends(get_db),
    _auth: dict = Depends(require_staff),
):
    """確定マッピングに従って患者を一括登録"""
    if mode not in ("split", "full_name"):
        raise HTTPException(400, "mode は split または full_name を指定してください")

    try:
        mapping: dict[str, int] = json.loads(mapping_json)
        mapping = {k: int(v) for k, v in mapping.items()}
    except (json.JSONDecodeError, ValueError):
        raise HTTPException(400, "mapping_json が不正です")

    # row_actions: [{row, action, patient_id?}, ...]
    row_actions: dict[int, dict] = {}
    if row_actions_json:
        try:
            actions_list = json.loads(row_actions_json)
            for a in actions_list:
                row_actions[int(a["row"])] = a
        except (json.JSONDecodeError, ValueError, KeyError):
            raise HTTPException(400, "row_actions_json が不正です")

    contents = await file.read()
    headers, data = _read_file(contents, file.filename or "import.csv")

    created_count = 0
    skipped_count = 0
    duplicate_count = 0
    adopted_count = 0
    updated_count = 0
    error_count = 0
    duplicates: list[dict] = []
    errors: list[dict] = []

    row_num = 0
    for row in data:
        row_num += 1
        if _is_empty_row(row):
            skipped_count += 1
            continue

        extracted = _extract_row(row, mapping)
        try:
            record = _build_patient_record(extracted, mode, row_num)
        except ValueError as e:
            error_count += 1
            errors.append({"row": row_num, "reason": str(e)})
            continue

        # 重複チェック
        dupes = await _find_duplicates(
            db,
            name=record["name"],
            phone=record.get("phone"),
            reading=record.get("reading"),
            birth=record.get("birth_date"),
        )
        if dupes:
            action = row_actions.get(row_num)
            if action:
                act = action.get("action", "skip")
                if act == "use_existing":
                    adopted_count += 1
                    continue
                elif act == "update_existing":
                    pid = action.get("patient_id")
                    if not pid:
                        error_count += 1
                        errors.append({"row": row_num, "reason": "update_existing に patient_id が必要です"})
                        continue
                    try:
                        await _update_existing_patient(db, int(pid), record)
                        updated_count += 1
                    except ValueError as e:
                        error_count += 1
                        errors.append({"row": row_num, "reason": str(e)})
                    continue
                else:
                    # skip or unknown
                    duplicate_count += 1
                    continue
            else:
                # アクション未指定 → 重複として報告
                duplicate_count += 1
                duplicates.append({
                    "row": row_num,
                    "data": {k: str(v) if v else "" for k, v in record.items()},
                    "candidates": dupes,
                })
                continue

        # 登録
        patient_number = await _generate_patient_number(db)
        patient = Patient(
            name=record["name"],
            registration_mode=mode,
            last_name=record.get("last_name"),
            middle_name=record.get("middle_name"),
            first_name=record.get("first_name"),
            reading=record.get("reading"),
            phone=record.get("phone"),
            birth_date=record.get("birth_date"),
            email=record.get("email"),
            notes=record.get("notes"),
            patient_number=patient_number,
        )
        db.add(patient)
        created_count += 1

    if created_count > 0 or updated_count > 0:
        await db.commit()

    total_rows = sum(1 for r in data if not _is_empty_row(r))

    return {
        "total_rows": total_rows,
        "created_count": created_count,
        "skipped_count": skipped_count,
        "duplicate_count": duplicate_count,
        "adopted_count": adopted_count,
        "updated_count": updated_count,
        "error_count": error_count,
        "duplicates": duplicates,
        "errors": errors,
    }


async def _update_existing_patient(
    db: AsyncSession, patient_id: int, record: dict[str, Any],
) -> None:
    """既存患者の限定フィールドを更新する（安全のため UPDATABLE_FIELDS のみ）"""
    result = await db.execute(
        select(Patient).where(Patient.id == patient_id, Patient.is_active == True)  # noqa: E712
    )
    patient = result.scalar_one_or_none()
    if not patient:
        raise ValueError(f"patient_id={patient_id} の患者が見つかりません")

    changed = False
    for field in UPDATABLE_FIELDS:
        new_val = record.get(field)
        if new_val is not None and new_val != "":
            current = getattr(patient, field, None)
            if current != new_val:
                setattr(patient, field, new_val)
                changed = True

    if not changed:
        return


def _build_patient_record(
    extracted: dict[str, str], mode: str, row_num: int,
) -> dict[str, Any]:
    """抽出データから Patient 用の dict を組み立てる"""
    record: dict[str, Any] = {}

    if mode == "split":
        last_name = _normalize_name(extracted.get("last_name", ""))
        first_name = _normalize_name(extracted.get("first_name", ""))
        middle_name = _normalize_name(extracted.get("middle_name", ""))

        # full_name 列のみマッピングで split モード指定 → 自動分割を試みる
        if not last_name and not first_name and extracted.get("full_name"):
            split_result = try_split_name(extracted["full_name"])
            if split_result:
                last_name, first_name = split_result
            else:
                raise ValueError(f"行{row_num}: 氏名を姓/名に分割できません。フルネームモードを使用してください")

        if not last_name:
            raise ValueError(f"行{row_num}: 姓が空です")
        if not first_name:
            raise ValueError(f"行{row_num}: 名が空です")
        record["last_name"] = last_name
        record["first_name"] = first_name
        record["middle_name"] = middle_name or None
        record["name"] = build_name("split", last_name, middle_name or None,
                                     first_name, None)
    else:
        full_name = _normalize_name(extracted.get("full_name", ""))
        if not full_name:
            raise ValueError(f"行{row_num}: フルネームが空です")
        record["name"] = full_name
        record["last_name"] = full_name
        record["first_name"] = None
        record["middle_name"] = None

    # 読み方
    reading = _normalize_name(extracted.get("reading", ""))
    record["reading"] = reading or None

    # 電話番号
    phone_raw = extracted.get("phone", "")
    record["phone"] = _normalize_phone(phone_raw)

    # 生年月日
    bd_raw = extracted.get("birth_date", "")
    if bd_raw:
        record["birth_date"] = _parse_date(bd_raw)
    else:
        record["birth_date"] = None

    # メール
    email = extracted.get("email", "").strip()
    record["email"] = email or None

    # 備考
    notes = extracted.get("notes", "").strip()
    record["notes"] = notes or None

    return record
